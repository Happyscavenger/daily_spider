# -*- coding:utf-8 -*-
from email.mime.multipart import MIMEMultipart
from email.utils import formataddr
from email.mime.application import MIMEApplication
import os, datetime, pymysql, xlrd, time, smtplib
from xlutils import copy
from  openpyxl import  Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active
class send_email():
	
	def __init__(self):
		self.sender = 'ayl1991923@qq.com'  # 发件人邮箱账号
		self.my_pass = 'xjgvtqstqkirbcad'  # 发件人授权码
		self.receivers = '807645038@qq.com' # 收件人邮箱账号，我这边发送给自己
		# self.receivers = ['673202141@qq.com', '1271396431@qq.com', '807645038@qq.com', '1135252170@qq.com']
		date_today = datetime.date.today()
		date_l = datetime.timedelta(days=1)
		self.date_y = date_today - date_l
		# table_name是我生成文件时(忽略)的命名，这里解释下
		self.name = '链家{}挂牌数据'.format(self.date_y) + '.xlsx'
		
		conn = pymysql.connect(
			host='rm-uf6t4r3u8vea8u3404o.mysql.rds.aliyuncs.com',
			port=3306,
			user='caijisa',
			passwd='Caijisa123',
			db='lansi_data_collection',
			charset='utf8'
		)
		self.cur = conn.cursor()
	
	def parse_sql(self):
	
		sql = """SELECT b.`Address`,a.`Area`,`Plate`,`Module`,a.`PropertyName`,a.`CompleteTime`,`UnitPrice`,`USD`,`Acreage`,`HouseTrait`,`Floor`,`AllFloor`,`RoomType`,`Shore`,`Fitment`,
			`Trade_id`,`URL`,CAST(`CreateDate` AS CHAR) AS `CreateDate` ,CAST(`UpdateDate` AS CHAR) AS `UpdateDate` ,`SetonStatus`,`latitudeX`,a.`LatitudeY`,
			CAST(`seton_data` AS CHAR) AS `seton_data`,`week_times`,`mouth_times`
			FROM `lansi_data_collection`.`ljseton` a INNER JOIN `lansi_data_collection`.`ljproperty` b ON a.`PropertyNo` = b.`PropertyNO` WHERE `UnitPrice`>10000 and a.`UpdateDate` >= '{}'""".format(
				self.date_y)
		self.cur.execute(sql)
		datas = self.cur.fetchall()
		return datas
	
	def write_excel_openpyxl(self,datas, filename):
		# 在内存创建一个工作簿obj
		result_wb = Workbook()
		# 第一个sheet是ws
		ws1 = result_wb.worksheets[0]
		# ws1=wb1.create_sheet('result',0)
		# 设置ws的名称
		ws1.title = "链家{}挂牌数据".format(self.date_y)
		row0 = ['地址', '区域', '板块', '环线', '小区', '年代', '单价', '总价', '面积', '房屋性质', '楼层', '总高', '户型', '朝向', '装修', '房源id',
		        'url',
		        '入库时间', 'UpdateDate', '状态',
		        'X坐标', 'Y坐标', '挂牌时间','周带看次数','月带看次数']
		ft = Font(name='Arial', size=11, bold=True)
		for k in range(len(row0)):
			ws1.cell(row=1, column=k + 1).value = row0[k]
			ws1.cell(row=1, column=k + 1).font = ft
		for i in range(1, len(datas) + 1):
			for j in range(1, len(row0) + 1):
				# col=get_column_letter(j)
				# ws1.cell('%s%s'%(col,i)).value='%s' % (data[j-1])
				ws1.cell(row=i + 1, column=j).value = datas[i - 1][j - 1]
		# 工作簿保存到磁盘
		result_wb.save(filename=filename)
	
	def mail(self):
		msg = MIMEMultipart()
		msg['From'] = formataddr(["杨力", self.sender])  # 括号里的对应发件人邮箱昵称、发件人邮箱账号
		msg['Subject'] = '链家{}挂牌数据'.format(self.date_y)  # 邮件的主题，也可以说是标题
		filepath = './' + self.name  # 绝对路径
		xlsxpart = MIMEApplication(open(filepath, 'rb').read())
		basename = self.name
		xlsxpart.add_header('Content-Disposition', 'attachment',
		                    filename=('gbk', '', basename))  # 注意：此处basename要转换为gbk编码，否则中文会有乱码。
		msg.attach(xlsxpart)
		server = smtplib.SMTP_SSL("smtp.qq.com", 465)  # 发件人邮箱中的SMTP服务器，端口是25
		server.login(self.sender, self.my_pass)  # 括号中对应的是发件人邮箱账号、邮箱授权码
		server.sendmail(self.sender, self.receivers, msg.as_string())  # 括号中对应的是发件人邮箱账号、收件人邮箱账号、发送邮件
		server.quit()  # 关闭连接
		os.remove(filepath)  # 删除文件
	
	def run(self):
		datas = self.parse_sql()
		self.write_excel_openpyxl(datas,self.name)
		self.mail()


if __name__ == '__main__':
	se = send_email()
	se.run()
